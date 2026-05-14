import openpyxl as xl
import os
from openpyxl.chart import BarChart, Reference


def update_price(filename, sheet_name, min_row, max_row, percent_discount, old_column, new_column):

    wb = xl.load_workbook(filename)
    sheet = wb[sheet_name]

    subtracted_price = 1.0 - percent_discount

    for row in range(min_row, max_row + 1):  # 2, 5
        cell = sheet.cell(row, old_column)
        corrected_price = float(cell.value) * subtracted_price
        corrected_price_cell = sheet.cell(row, new_column)
        corrected_price_cell.value = corrected_price

#    values = Reference(sheet,
#                       min_row=2,
#                       max_row=sheet.max_row,
#                       min_col=4,
#                       max_col=4)
#
#    chart = BarChart()
#    chart.add_data(values)
#    sheet.add_chart(chart, 'e2')

    wb.save(filename)


update_price('transactions.xlsx', 'Sheet1', 2, 4, 0.1, 2, 3)
