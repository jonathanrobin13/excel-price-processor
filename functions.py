import openpyxl as xl
from openpyxl.chart import BarChart, Reference


class Discount:

    def __init__(self, filename, sheet_name, min_row, max_row, old_column, new_column):
        self.filename = filename
        self.sheet_name = sheet_name
        self.min_row = min_row
        self.max_row = max_row
        self.old_column = old_column
        self.new_column = new_column

    def percentage_discount(self, percent_discount):

        wb = xl.load_workbook(self.filename)
        sheet = wb[self.sheet_name]

        subtracted_price = 1.0 - float(percent_discount)

        # Goes through each cell and creates a new cell with the discounted value
        for row in range(self.min_row, self.max_row + 1):
            cell = sheet.cell(row, self.old_column)
            corrected_price = float(cell.value) * subtracted_price
            corrected_price_cell = sheet.cell(row, self.new_column)
            corrected_price_cell.value = corrected_price

        wb.save(self.filename)

    def fixed_discount(self, fixed_discount):

        wb = xl.load_workbook(self.filename)
        sheet = wb[self.sheet_name]

        for row in range(self.min_row, self.max_row + 1):
            cell = sheet.cell(row, self.old_column)
            corrected_price = float(cell.value) - fixed_discount
            new_cell = sheet.cell(row, self.new_column)
            new_cell.value = corrected_price

        wb.save(self.filename)


def column_finder(column):
    alphabet_list = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K',
                     'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']

    column = column.capitalize()

    # adds 1 because the index starts at 0 but the column starts at one
    column_number = alphabet_list.index(column) + 1
    return column_number


# a min price and a max price finder for later


# Different Function for later
#    values = Reference(sheet,
#                       min_row=2,
#                       max_row=sheet.max_row,
#                       min_col=4,
#                       max_col=4)
#
#    chart = BarChart()
#    chart.add_data(values)
#    sheet.add_chart(chart, 'e2')
