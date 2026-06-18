import openpyxl as xl
from openpyxl.chart import BarChart, Reference


class Price:

    def __init__(self, filename, sheet_name, min_row, max_row, original_price_col, discount_price_col, product_name_col):
        self.filename = filename
        self.sheet_name = sheet_name
        self.min_row = min_row
        self.max_row = max_row
        self.original_price_col = original_price_col
        self.discount_price_col = discount_price_col
        self.product_name_col = product_name_col

    def percentage_discount(self, percent_discount):

        wb = xl.load_workbook(self.filename)
        sheet = wb[self.sheet_name]

        # finds out how much percent is left of the original price once the discount is applied
        subtracted_price = 1.0 - float(percent_discount)

        # adds one so that it starts at min row and ends at max row
        for row in range(self.min_row, self.max_row + 1):
            cell = sheet.cell(row, self.original_price_col)
            corrected_price = float(cell.value) * subtracted_price
            corrected_price_cell = sheet.cell(row, self.discount_price_col)
            corrected_price_cell.value = corrected_price

        wb.save(self.filename)

    def fixed_discount(self, fixed_discount):

        wb = xl.load_workbook(self.filename)
        sheet = wb[self.sheet_name]

        # adds one so that it starts at min row and ends at max row
        for row in range(self.min_row, self.max_row + 1):
            cell = sheet.cell(row, self.original_price_col)
            corrected_price = float(cell.value) - fixed_discount
            new_cell = sheet.cell(row, self.discount_price_col)
            new_cell.value = corrected_price

        wb.save(self.filename)


# decision finds whether the user wants to find min or max

    def min_max_finder(self, decision):

        wb = xl.load_workbook(self.filename)
        sheet = wb[self.sheet_name]

        min = 0
        min_product_name = ""
        max = 0
        max_product_name = ""

        if decision == 0:  # if user asks for min

            for row in range(self.min_row, self.max_row + 1):
                cell = sheet.cell(row, self.original_price_col)

                # if this is the first iteration, make the first value as the minimum
                if row == self.min_row:
                    min = cell.value
                    # Makes the product name for the minimum value
                    min_product_cell = sheet.cell(row, self.product_name_col)
                    min_product_name = min_product_cell.value

                else:
                    # sets the new min
                    new_min = min(min, cell.value)

                    if new_min == cell.value:
                        min = cell.value
                    min_product_cell = sheet.cell(row, self.product_name_col)
                    min_product_name = min_product_cell.value

            return min, min_product_name

        # If user asks for max
        elif decision == 1:

            for row in range(self.min_row, self.max_row + 1):
                cell = sheet.cell(row, self.original_price_col)

                # if this is the first iteration, make the first value as the maximum
                if row == self.min_row:
                    max = cell.value
                    # Makes the product name for the maximum value
                    max_product_cell = sheet.cell(row, self.product_name_col)
                    max_product_name = max_product_cell.value

                else:
                    # sets the new max
                    new_max = max(max, cell.value)

                    if new_max == cell.value:
                        max = cell.value
                    max_product_cell = sheet.cell(row, self.product_name_col)
                    max_product_name = max_product_cell.value

            return max, max_product_name


def column_finder(column):
    alphabet_list = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K',
                     'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']

    column = column.capitalize()

    # adds 1 because the index starts at 0 but the column starts at one
    column_number = alphabet_list.index(column) + 1
    return column_number


# a min price and a max price finder for later


# Different Function for later
#    values = Reference(sheet,
#                       min_row=2,
#                       max_row=sheet.max_row,
#                       min_col=4,
#                       max_col=4)
#
#    chart = BarChart()
#    chart.add_data(values)
#    sheet.add_chart(chart, 'e2')
